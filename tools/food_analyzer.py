#!/usr/bin/env python3
"""
NutriLens AI 음식 분석 엔진 (2단계 핵심 도구)
─────────────────────────────────────────────
사진 한 장 → AI가 음식 인식 → DB에서 영양소 매칭 → 결과 반환

사용법:
  python food_analyzer.py --image 사진경로.jpg
  python food_analyzer.py --image 사진경로.jpg --api-key YOUR_OPENAI_KEY

환경변수:
  OPENAI_API_KEY: OpenAI API 키 (.env 파일에 저장)
"""

import os
import sys
import json
import base64
import argparse
from pathlib import Path

# ── .env 파일 로드 ──
def load_env():
    env_paths = [
        Path(__file__).parent.parent / '.env',
        Path.cwd() / '.env',
    ]
    for env_path in env_paths:
        if env_path.exists():
            with open(env_path) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#') and '=' in line:
                        key, val = line.split('=', 1)
                        os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))
            break

load_env()

# ══════════════════════════════════════════════════════════════════
#  GI(혈당지수) / GL(혈당부하) 엔진  —  설계서 132 · 검증패킷 133
# ──────────────────────────────────────────────────────────────────
#  3계층:
#    T1 DB 조회   : food_id → gi_table_v1.csv 의 gi (결정론 · 정답 경로)
#    T2 프록시    : gi 없음 → 가용탄수 게이트 / 카테고리 기본 GI
#    T3 AI 추정   : DB·카테고리 모두 미상 → AI 값 사용 + 낮은 confidence 플래그
#  GI는 세기(intensive) → 양 비율을 곱하지 않는다. 양은 GL에서만 반영.
# ══════════════════════════════════════════════════════════════════

GI_FIELDS = ('gi', 'gi_source', 'gi_confidence', 'gi_ref')

# T2 프록시: 가용탄수가 이 값 미만이면 GI 무의미 → GL≈0 (설계서 §3)
LOW_CARB_GATE_G = 10.0

# T2 프록시: subcategory → 기본 GI (직접 측정치가 없을 때만 사용)
CATEGORY_GI = {
    # 곡류·주식
    '밥류': 70, '죽류': 85, '면류': 55, '분식류': 70, '정식': 68, '도시락': 70,
    '빵류': 72, '파스타': 50, '탄수화물': 55,
    # 단백질·저탄수 (탄수는 대개 양념 유래)
    '육류': 60, '해산물': 55, '찜류': 60, '전골류': 60, '전류': 65,
    '탕류': 60, '찌개류': 55, '국류': 55, '반찬류': 55, '양념류': 60,
    '고단백_저지방': 40, '고단백_중지방': 40, '고단백_고지방': 40, '식물성단백': 20,
    # 외국·프랜차이즈
    '일식': 60, '중식': 62, '동남아': 62, '양식': 62, '브런치': 65, '인도': 65,
    '버거': 66, '샌드위치': 60, '치킨': 60, '피자': 60, '사이드': 60, '기타': 60,
    # 간식·음료·유제품
    '샐러드': 45, '과일': 45, '유제품': 35, '우유': 40, '커피': 40, '차': 40,
    '주스': 50, '탄산음료': 63, '음료': 60, '디저트': 65, '과자': 65,
    '간식': 60, '보충제': 50,
}
CATEGORY_GI_FALLBACK = {
    'korean': 60, 'foreign_popular': 60, 'franchise': 65,
    'snack_drink': 60, 'diet_fitness': 50,
}

GL_LOW_MAX = 10      # ≤10 low
GL_MED_MAX = 19      # 11~19 med, ≥20 high


def _gi_table_paths():
    """GI 표 탐색 경로: **IP 원본 우선** → 코드 저장소 사본 (원칙3).

    원칙3 은 "IP 가 정본, 코드 저장소엔 사본만" 이다. 따라서 정본이 있으면 정본을 읽어야 한다.
    (세션30 정정) 이전 구현은 사본을 **먼저** 봤다 — IP 원본만 고치면 엔진이 낡은 사본을 계속
    쓰면서도 조용히 성공하는 침묵 함정이었다. 두 파일이 어긋나면 아래에서 소리내어 경고한다.
    사본은 IP 가 동봉되지 않는 배포 환경을 위한 폴백이다.
    """
    here = Path(__file__).parent
    return [
        here.parent / 'IP' / 'content' / 'gi_table_v1.csv',   # ← 정본
        here / 'data' / 'gi_table_v1.csv',                    # ← 사본(배포 폴백)
    ]


def _warn_gi_table_drift():
    """정본과 사본이 어긋나면 경고한다. 조용히 한쪽을 고르지 않는다.

    표류를 CI 에서 잡으려면: python3 tools/build_gi_table_ts.py --check
    """
    import hashlib

    canon, copy = _gi_table_paths()
    if not (canon.exists() and copy.exists()):
        return
    def norm(p):   # CRLF/LF 차이는 표류가 아니다
        return hashlib.sha256(p.read_bytes().replace(b'\r\n', b'\n')).hexdigest()
    if norm(canon) != norm(copy):
        import warnings
        warnings.warn(
            f"GI 표 표류 감지: 정본({canon})과 사본({copy})의 내용이 다릅니다. "
            f"정본을 읽습니다. 사본 갱신: python3 tools/build_gi_table_ts.py",
            RuntimeWarning, stacklevel=2,
        )


def load_gi_table(path=None):
    """gi_table_v1.csv → {food_id: {gi, gi_source, gi_confidence, gi_ref}}"""
    import csv

    if path is None:
        _warn_gi_table_drift()
    candidates = [Path(path)] if path else _gi_table_paths()
    for p in candidates:
        if not p.exists():
            continue
        table = {}
        with open(p, encoding='utf-8-sig', newline='') as f:
            for row in csv.DictReader(f):
                fid = (row.get('food_id') or '').strip()
                raw_gi = (row.get('gi') or '').strip()
                if not fid or not raw_gi:
                    continue          # 공란 = 미상 → T2에 위임
                try:
                    gi = int(round(float(raw_gi)))
                except ValueError:
                    continue
                if not 0 <= gi <= 110:
                    continue
                table[fid] = {
                    'gi': gi,
                    'gi_source': (row.get('gi_source') or 'db_measured').strip(),
                    'gi_confidence': (row.get('gi_confidence') or 'low').strip(),
                    'gi_ref': (row.get('gi_ref') or '').strip(),
                }
        return table
    return {}


def available_carb(carbs_g, fiber_g):
    """가용탄수 = 탄수 − 식이섬유 (음수면 0)"""
    try:
        return max(float(carbs_g or 0) - float(fiber_g or 0), 0.0)
    except (TypeError, ValueError):
        return 0.0


def gl_category(gl):
    """GL 등급 (당뇨병 표준): ≤10 low · 11~19 med · ≥20 high"""
    if gl <= GL_LOW_MAX:
        return 'low'
    if gl <= GL_MED_MAX:
        return 'med'
    return 'high'


def resolve_gi(food, db_match=None, gi_table=None):
    """
    한 음식의 GI를 3계층으로 해석하여 (gi, source, confidence, ref) 반환.
    GI는 세기 → 양 비율 미적용.
    """
    gi_table = gi_table if gi_table is not None else {}

    # ── T1: DB(food_id) 조회 — 결정론 정답 경로 ──
    food_id = (db_match or {}).get('food_id') or food.get('db_food_id')
    if food_id and food_id in gi_table:
        e = gi_table[food_id]
        return e['gi'], e['gi_source'], e['gi_confidence'], e['gi_ref']

    # ── T2: 가용탄수 게이트 → 카테고리 프록시 ──
    src = db_match or food
    avail = available_carb(src.get('carbs_g'), src.get('fiber_g'))
    if avail < LOW_CARB_GATE_G:
        # 가용탄수 ≈ 0 → GI 무의미(고기·기름·채소·저탄수 찌개). GL≈0.
        return None, 'low_carb', 'na', 'T2: 가용탄수 %.1fg < %.0fg → GI 무의미' % (avail, LOW_CARB_GATE_G)

    subcat = (src.get('subcategory') or '').strip()
    cat = (src.get('category') or '').strip()
    if subcat in CATEGORY_GI:
        return CATEGORY_GI[subcat], 'category', 'low', 'T2: subcategory=%s 기본 GI' % subcat
    if cat in CATEGORY_GI_FALLBACK:
        return CATEGORY_GI_FALLBACK[cat], 'category', 'low', 'T2: category=%s 기본 GI' % cat

    # ── T3: AI 추정값이 있으면 사용(최후 · 항상 low + 플래그) ──
    ai_gi = food.get('gi')
    if isinstance(ai_gi, (int, float)) and 0 <= ai_gi <= 110:
        return int(round(ai_gi)), 'ai_estimate', 'low', 'T3: AI 추정 — 검수 필요'

    return None, 'unknown', 'na', 'T1~T3 모두 미상'


def compute_glycemic(analysis, gi_table=None):
    """
    설계서 §4 결정론 공식으로 음식별 GL + 식사 GL/식사 GI 산출.
      가용탄수 = carbs_g − fiber_g (음수면 0)
      음식 GL  = gi × 가용탄수 / 100
      식사 GL  = Σ(음식 GL)
      식사 GI  = Σ(gi × 가용탄수) / Σ(가용탄수)   (탄수 가중평균)
    AI 추론 0. 엔진이 이미 가진 값만으로 계산.
    """
    if 'error' in analysis or 'foods' not in analysis:
        return analysis

    gi_table = gi_table if gi_table is not None else {}

    total_gl = 0.0
    weighted_gi = 0.0
    total_avail = 0.0

    for food in analysis['foods']:
        if not food.get('gi_source'):
            gi, src, conf, ref = resolve_gi(food, None, gi_table)
            food['gi'] = gi
            food['gi_source'] = src
            food['gi_confidence'] = conf
            food['gi_ref'] = ref

        avail = available_carb(food.get('carbs_g'), food.get('fiber_g'))
        gi = food.get('gi')
        gl = round((gi or 0) * avail / 100, 1)

        food['available_carb_g'] = round(avail, 1)
        food['gl'] = gl
        food['gl_category'] = gl_category(gl)

        total_gl += gl
        if gi is not None:
            weighted_gi += gi * avail
            total_avail += avail

    summary = analysis.setdefault('meal_summary', {})
    meal_gl = round(total_gl, 1)
    summary['meal_gl'] = meal_gl
    summary['meal_gl_category'] = gl_category(meal_gl)
    summary['meal_gi'] = round(weighted_gi / total_avail, 1) if total_avail > 0 else None
    summary['meal_available_carb_g'] = round(total_avail, 1)
    # 정직한 불확실성(설계서 §2-5): 코치는 단정 금지
    confs = [f.get('gi_confidence') for f in analysis['foods'] if f.get('gi') is not None]
    summary['gi_confidence_min'] = (
        'low' if 'low' in confs else 'med' if 'med' in confs else 'high' if confs else 'na'
    )
    return analysis


# ── 음식 DB 로드 ──
def load_food_db(db_path=None, gi_table_path=None):
    """엑셀 DB를 로드하여 딕셔너리 리스트로 반환 (+ gi_table_v1.csv 머지)"""
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 필요: pip install openpyxl --break-system-packages")
        sys.exit(1)

    if db_path is None:
        db_path = Path(__file__).parent.parent / 'NutriLens_음식DB.xlsx'

    if not Path(db_path).exists():
        print(f"DB 파일을 찾을 수 없습니다: {db_path}")
        return []

    wb = openpyxl.load_workbook(db_path, read_only=True, data_only=True)
    ws = wb['음식DB_전체']

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    foods = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            food = dict(zip(headers, row))
            foods.append(food)
    wb.close()

    # GI 4컬럼 머지 (xlsx에 gi 컬럼이 없어도 CSV 원본에서 붙인다)
    gi_table = load_gi_table(gi_table_path)
    for f in foods:
        entry = gi_table.get(f.get('food_id'))
        for k in GI_FIELDS:
            f[k] = entry[k] if entry else None

    return foods


def build_food_list_for_prompt(foods):
    """DB 음식 목록을 프롬프트에 넣을 간결한 텍스트로 변환"""
    categories = {}
    for f in foods:
        cat = f.get('category', 'other')
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(f['name_ko'])

    lines = []
    cat_names = {
        'korean': '한식',
        'diet_fitness': '다이어트/피트니스',
        'foreign_popular': '외국음식',
        'franchise': '프랜차이즈',
        'snack_drink': '간식/음료',
    }
    for cat, names in categories.items():
        label = cat_names.get(cat, cat)
        lines.append(f"[{label}] {', '.join(names[:50])}")  # 카테고리당 최대 50개
    return '\n'.join(lines)


# ── AI 분석 프롬프트 ──
SYSTEM_PROMPT = """당신은 NutriLens의 AI 음식 영양 분석 전문가입니다.

## 역할
사진에 보이는 모든 음식을 정확하게 식별하고, 각 음식의 영양 성분을 분석합니다.

## 규칙
1. 사진에 보이는 음식을 **모두** 개별적으로 식별하세요 (메인, 사이드, 음료 포함)
2. 각 음식의 **예상 양(g)**을 추정하세요 (접시/그릇 크기 참고)
3. 영양소는 추정 양 기준으로 계산하세요
4. 한국 음식이면 category를 "korean", 다이어트 식단이면 "diet_fitness", 외국 음식이면 "foreign_popular", 프랜차이즈면 "franchise", 간식/음료면 "snack_drink"으로 분류
5. confidence는 0.0~1.0 (1.0 = 100% 확신)
6. 프랜차이즈 메뉴가 보이면 brand 필드에 브랜드명 포함
7. 반드시 아래 JSON 형식으로만 답변하세요. 다른 텍스트 없이 JSON만 출력하세요.

## 응답 형식
```json
{
  "foods": [
    {
      "name_ko": "음식 한국어 이름",
      "name_en": "English Name",
      "category": "korean",
      "subcategory": "밥류",
      "estimated_serving_g": 300,
      "calories_kcal": 450,
      "protein_g": 15,
      "carbs_g": 65,
      "fat_g": 12,
      "fiber_g": 2,
      "sodium_mg": 800,
      "sugar_g": 3,
      "confidence": 0.92,
      "brand": "",
      "cooking_method": "볶음",
      "tags": "한끼,매운맛"
    }
  ],
  "meal_summary": {
    "total_calories": 450,
    "total_protein": 15,
    "total_carbs": 65,
    "total_fat": 12,
    "meal_type": "점심",
    "health_score": 7,
    "one_line_comment": "균형 잡힌 한끼입니다. 식이섬유를 위해 채소를 추가하면 더 좋아요."
  }
}
```

## health_score 기준 (1~10)
- 9~10: 영양 균형 우수, 저나트륨, 적정 칼로리
- 7~8: 대체로 양호, 약간의 개선 여지
- 5~6: 보통, 특정 영양소 과다/부족
- 3~4: 영양 불균형, 고칼로리/고나트륨
- 1~2: 매우 불균형, 과도한 열량/지방/나트륨"""


def encode_image(image_path):
    """이미지를 base64로 인코딩"""
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def analyze_food_image(image_path, api_key=None, model="gpt-4o"):
    """
    음식 사진을 GPT-4o Vision으로 분석

    Args:
        image_path: 이미지 파일 경로
        api_key: OpenAI API 키 (없으면 환경변수에서 가져옴)
        model: 사용할 모델 (기본: gpt-4o)

    Returns:
        dict: 분석 결과 (JSON)
    """
    try:
        import httpx
    except ImportError:
        try:
            import requests as httpx
            httpx.Client = None  # fallback marker
        except ImportError:
            print("httpx 또는 requests 필요: pip install httpx --break-system-packages")
            sys.exit(1)

    if api_key is None:
        api_key = os.environ.get('OPENAI_API_KEY')

    if not api_key:
        return {
            "error": "OPENAI_API_KEY가 설정되지 않았습니다.",
            "help": "1) .env 파일에 OPENAI_API_KEY=sk-... 추가하거나\n"
                    "2) --api-key 옵션으로 전달하세요.\n"
                    "3) OpenAI API 키는 https://platform.openai.com/api-keys 에서 발급"
        }

    if not Path(image_path).exists():
        return {"error": f"이미지 파일을 찾을 수 없습니다: {image_path}"}

    # 이미지 인코딩
    base64_image = encode_image(image_path)
    ext = Path(image_path).suffix.lower()
    media_type = {
        '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
        '.png': 'image/png', '.gif': 'image/gif',
        '.webp': 'image/webp',
    }.get(ext, 'image/jpeg')

    # API 호출
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }

    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": "이 사진에 있는 음식을 분석해주세요. 모든 음식을 개별적으로 식별하고 영양 성분을 알려주세요."
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{media_type};base64,{base64_image}",
                            "detail": "high"
                        }
                    }
                ]
            }
        ],
        "max_tokens": 2000,
        "temperature": 0.2,
    }

    try:
        import httpx
        response = httpx.post(
            "https://api.openai.com/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=30.0,
        )
        result = response.json()
    except Exception as e:
        return {"error": f"API 호출 실패: {str(e)}"}

    if "error" in result:
        return {"error": f"OpenAI API 에러: {result['error'].get('message', str(result['error']))}"}

    # 응답 파싱
    content = result["choices"][0]["message"]["content"]

    # JSON 추출 (```json ... ``` 블록이 있을 수 있음)
    if "```json" in content:
        content = content.split("```json")[1].split("```")[0]
    elif "```" in content:
        content = content.split("```")[1].split("```")[0]

    try:
        analysis = json.loads(content.strip())
    except json.JSONDecodeError:
        return {
            "error": "AI 응답을 파싱할 수 없습니다",
            "raw_response": content
        }

    return analysis


def match_with_db(analysis, foods_db):
    """
    AI 분석 결과를 DB와 매칭하여 보정
    AI가 인식한 음식이 DB에 있으면 DB의 공인 데이터를 사용
    """
    if "error" in analysis or "foods" not in analysis:
        return analysis

    # food_id → GI 4컬럼 (load_food_db가 이미 머지해둔 값을 인덱싱)
    gi_table = {
        db_food['food_id']: {k: db_food.get(k) for k in GI_FIELDS}
        for db_food in foods_db
        if db_food.get('food_id') and db_food.get('gi') is not None
    }

    for food in analysis["foods"]:
        ai_name = food.get("name_ko", "")

        # DB에서 이름으로 검색 (정확 매칭 → 부분 매칭)
        exact_match = None
        partial_matches = []

        for db_food in foods_db:
            db_name = db_food.get("name_ko", "")
            if ai_name == db_name:
                exact_match = db_food
                break
            elif ai_name in db_name or db_name in ai_name:
                partial_matches.append(db_food)

        match = exact_match or (partial_matches[0] if partial_matches else None)

        if match:
            # DB 매칭 성공: 1인분 영양소를 AI 추정 양에 비례하여 보정
            db_serving = match.get('serving_size_g', 100) or 100
            ai_serving = food.get('estimated_serving_g', db_serving) or db_serving
            ratio = ai_serving / db_serving

            food['db_matched'] = True
            food['db_food_id'] = match.get('food_id', '')
            food['db_name'] = match.get('name_ko', '')

            # DB 값으로 보정 (양 비율 적용)
            for field_pair in [
                ('calories_kcal', 'calories_kcal'),
                ('protein_g', 'protein_g'),
                ('carbs_g', 'carbs_g'),
                ('fat_g', 'fat_g'),
                ('fiber_g', 'fiber_g'),
                ('sodium_mg', 'sodium_mg'),
                ('sugar_g', 'sugar_g'),
            ]:
                ai_field, db_field = field_pair
                db_val = match.get(db_field)
                if db_val is not None and isinstance(db_val, (int, float)):
                    food[ai_field] = round(db_val * ratio, 1)

            food['source'] = 'DB_MATCHED'
        else:
            food['db_matched'] = False
            food['source'] = 'AI_ESTIMATED'

        # ── GI 부착 (T1 DB조회 / T2 프록시 / T3 AI추정) ──
        # 주의: GI는 세기(intensive) → 양 비율(ratio)을 곱하지 않는다. 양은 GL에서만 반영.
        gi, gi_src, gi_conf, gi_ref = resolve_gi(food, match, gi_table)
        food['gi'] = gi
        food['gi_source'] = gi_src
        food['gi_confidence'] = gi_conf
        food['gi_ref'] = gi_ref

    # meal_summary 재계산
    if "meal_summary" in analysis:
        analysis["meal_summary"]["total_calories"] = round(sum(f.get("calories_kcal", 0) for f in analysis["foods"]), 1)
        analysis["meal_summary"]["total_protein"] = round(sum(f.get("protein_g", 0) for f in analysis["foods"]), 1)
        analysis["meal_summary"]["total_carbs"] = round(sum(f.get("carbs_g", 0) for f in analysis["foods"]), 1)
        analysis["meal_summary"]["total_fat"] = round(sum(f.get("fat_g", 0) for f in analysis["foods"]), 1)

    # ── GL 계산(음식별 + 식사 합산) — 결정론, AI 추론 0 ──
    compute_glycemic(analysis, gi_table)

    return analysis


def format_result(analysis):
    """분석 결과를 사람이 읽기 좋은 텍스트로 변환"""
    if "error" in analysis:
        return f"오류: {analysis['error']}"

    lines = []
    lines.append("=" * 50)
    lines.append("  NutriLens 음식 분석 결과")
    lines.append("=" * 50)

    for i, food in enumerate(analysis.get("foods", []), 1):
        matched = "DB" if food.get('db_matched') else "AI추정"
        conf = food.get('confidence', 0)
        lines.append(f"\n[{i}] {food['name_ko']} ({food.get('name_en', '')})  [{matched}] 확신도:{conf:.0%}")
        lines.append(f"    양: {food.get('estimated_serving_g', '?')}g")
        lines.append(f"    칼로리: {food.get('calories_kcal', '?')} kcal")
        lines.append(f"    단백질: {food.get('protein_g', '?')}g  |  탄수화물: {food.get('carbs_g', '?')}g  |  지방: {food.get('fat_g', '?')}g")
        lines.append(f"    식이섬유: {food.get('fiber_g', '?')}g  |  나트륨: {food.get('sodium_mg', '?')}mg  |  당류: {food.get('sugar_g', '?')}g")
        gi = food.get('gi')
        if gi is not None:
            lines.append(f"    GI: {gi} ({food.get('gi_source', '?')}/{food.get('gi_confidence', '?')})"
                         f"  |  GL: {food.get('gl', '?')} [{food.get('gl_category', '?')}]")
        elif food.get('gi_source'):
            lines.append(f"    GI: — (가용탄수 낮음 → GI 무의미)  |  GL: {food.get('gl', 0)} [low]")

    summary = analysis.get("meal_summary", {})
    if summary:
        lines.append(f"\n{'─' * 50}")
        lines.append(f"  총 칼로리: {summary.get('total_calories', '?')} kcal")
        lines.append(f"  단백질 {summary.get('total_protein', '?')}g / 탄수 {summary.get('total_carbs', '?')}g / 지방 {summary.get('total_fat', '?')}g")
        if summary.get('meal_gl') is not None:
            lines.append(f"  식사 GL: {summary['meal_gl']} [{summary.get('meal_gl_category', '?')}]"
                         f"  |  식사 GI(참고): {summary.get('meal_gi', '—')}"
                         f"  |  GI 신뢰도: {summary.get('gi_confidence_min', '?')}")
        lines.append(f"  식사 유형: {summary.get('meal_type', '?')}  |  건강 점수: {summary.get('health_score', '?')}/10")
        lines.append(f"\n  AI 코멘트: {summary.get('one_line_comment', '')}")
    lines.append("=" * 50)

    return '\n'.join(lines)


# ── 메인 실행 ──
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="NutriLens AI 음식 분석기")
    parser.add_argument("--image", "-i", required=True, help="분석할 음식 사진 경로")
    parser.add_argument("--api-key", "-k", default=None, help="OpenAI API 키")
    parser.add_argument("--model", "-m", default="gpt-4o", help="사용할 모델 (기본: gpt-4o)")
    parser.add_argument("--json", action="store_true", help="JSON 형식으로 출력")
    parser.add_argument("--no-db", action="store_true", help="DB 매칭 건너뛰기")
    args = parser.parse_args()

    print(f"이미지 분석 중: {args.image}")
    print("GPT-4o Vision API에 요청 중...")

    # 1. AI 분석
    analysis = analyze_food_image(args.image, api_key=args.api_key, model=args.model)

    # 2. DB 매칭 (선택)
    if not args.no_db and "error" not in analysis:
        print("DB 매칭 중...")
        foods_db = load_food_db()
        if foods_db:
            analysis = match_with_db(analysis, foods_db)
            print(f"DB 매칭 완료 ({len(foods_db)}종 DB 사용)")

    # 3. 출력
    if args.json:
        print(json.dumps(analysis, ensure_ascii=False, indent=2))
    else:
        print(format_result(analysis))
